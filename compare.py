from openpyxl import load_workbook

def fill_column_based_on_time(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active

    # 读取第7列（G）为键，第8列（H）为值，组成一个字典
    time_data_map = {}
    for row in range(2, ws.max_row + 1):
        time_val = ws.cell(row=row, column=7).value
        data_val = ws.cell(row=row, column=8).value
        if time_val:
            time_data_map[time_val] = data_val

    # 遍历第1列（A），将匹配的数据写入第9列（I）
    for row in range(2, ws.max_row + 1):
        time_val = ws.cell(row=row, column=1).value
        if time_val in time_data_map:
            ws.cell(row=row, column=9).value = time_data_map[time_val]

    wb.save(excel_path)

if __name__ == "__main__":
    path = "C:\\Users\\hanji\\Desktop\\tmp\\湖泊面积\\逐月面积-对比.xlsx"

    fill_column_based_on_time(path)