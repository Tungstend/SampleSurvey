import pandas as pd
import math
from openpyxl import load_workbook

def get_time(input_path):
    df = pd.read_excel(input_path)
    time_col = df.iloc[:, 0].to_numpy()
    return time_col

def convert_number_to_year_month(num):
    year = int(num)
    month_fraction = num - year
    month = math.ceil(month_fraction / (1/12))  # 多加1月
    if month > 12:
        year += 1
        month = 1
    return f"{year}年{month}月"

def write_time_to_excel(numbers, excel_path):
    # 加载或创建工作簿
    try:
        wb = load_workbook(excel_path)
    except FileNotFoundError:
        from openpyxl import Workbook
        wb = Workbook()
    ws = wb.active

    # 写入转换后的时间
    for i, num in enumerate(numbers, start=1):
        date_str = convert_number_to_year_month(num)
        ws.cell(row=i, column=1, value=date_str)

    # 保存文件
    wb.save(excel_path)

def generate_time(input_path, output_path):
    write_time_to_excel(get_time(input_path), output_path)

if __name__ == "__main__":
    input_path = "C:\\Users\\hanji\\Desktop\\tmp\\湖泊面积\\zzzz\\input.xlsx"
    output_path = "C:\\Users\\hanji\\Desktop\\tmp\\湖泊面积\\zzzz\\output.xlsx"

    generate_time(input_path, output_path)