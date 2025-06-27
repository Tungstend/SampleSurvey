import os
import re
import pandas as pd
import warnings
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import coordinate_from_string

warnings.simplefilter("ignore", UserWarning)

def is_number(value):
    try:
        float(value)
        return True
    except:
        return False

def parse_meter_string(s):
    try:
        return float(s.strip().replace("m", ""))
    except:
        return s

def read_fixed_cells(sheet, cells):
    values = []
    for cell in cells:
        val = parse_meter_string(sheet[cell].value)
        values.append(float(val) if is_number(val) else None)
    return values

def read_keyword_cells(sheet, config, search_range):
    results = {field: None for field in config}
    start_coord, end_coord = search_range
    start_col_letter, start_row = coordinate_from_string(start_coord)
    end_col_letter, end_row = coordinate_from_string(end_coord)
    start_col = column_index_from_string(start_col_letter)
    end_col = column_index_from_string(end_col_letter)

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            value = str(cell.value).strip() if cell.value else ""
            for field, rule in config.items():
                if results[field] is not None:
                    continue
                if any(alias in value for alias in rule["aliases"]):
                    offset_row, offset_col = rule["offset"]
                    data_cell = sheet.cell(row=row + offset_row, column=col + offset_col)
                    data_value = data_cell.value
                    results[field] = float(data_value) if is_number(data_value) else None
    return [results[field] for field in config]

def parse_filename(filename):
    name = os.path.splitext(filename)[0]
    match = re.match(r'^Y(\d{2})(\d{2})-(\d+)', name)
    if not match:
        return name, 99, 99, 999999
    month = int(match.group(1))
    day = int(match.group(2))
    main = int(match.group(3))
    return name, month, day, main

# ====== 主流程：读取所有 Excel 数据 ======

def process_excel_folder(folder_path, fixed_cells, keyword_config, search_range):
    records = []
    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            full_path = os.path.join(folder_path, filename)
            name, month, day, main = parse_filename(filename)
            try:
                wb = load_workbook(full_path, data_only=True)
                sheet = wb.active
                fixed_values = read_fixed_cells(sheet, fixed_cells)
                keyword_values = read_keyword_cells(sheet, keyword_config, search_range)
                records.append((name, *fixed_values, *keyword_values, month, day, main))
            except Exception as e:
                print(f"[异常] 无法读取文件: {filename}，原因: {e}")
    fixed_headers = fixed_cells
    keyword_headers = list(keyword_config.keys())
    headers = ["编号"] + fixed_headers + keyword_headers
    df = pd.DataFrame(records, columns=headers + ["月", "日", "主编号"])
    return df

def preprocess_template_expand_sub_ids(template_path, data_df):
    wb = load_workbook(template_path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=False))  # 返回单元格对象
    header = [cell.value for cell in rows[0]]
    point_index = {
        str(ws.cell(row=i+2, column=1).value): i+2
        for i in range(1, len(rows)-1)
        if ws.cell(row=i+2, column=1).value is not None
    }

    # 构建数据中子编号映射 {主编号: [子编号列表]}
    sub_map = {}
    for fid in data_df["编号"]:
        parts = str(fid).split("-")
        if len(parts) >= 3:
            base = "-".join(parts[:2])
            sub_map.setdefault(base, []).append(str(fid))

    for base_id, sub_ids in sub_map.items():
        if base_id in point_index:
            row_idx = point_index[base_id]
            # 记录原行数据
            old_row_data = [ws.cell(row=row_idx, column=col).value for col in range(1, ws.max_column + 1)]
            # 删除原主编号行
            ws.delete_rows(row_idx, 1)
            # 插入子编号行，复制原行数据
            for offset, sub_id in enumerate(sorted(sub_ids)):
                insert_row = row_idx + offset
                ws.insert_rows(insert_row)
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=insert_row, column=col).value = old_row_data[col - 1]
                ws.cell(row=insert_row, column=1).value = sub_id  # 更新编号列

            # 更新 point_index 行号引用（后续数据写入使用）
            for pid in list(point_index):
                if point_index[pid] > row_idx:
                    point_index[pid] += len(sub_ids) - 1

    temp_path = template_path.replace(".xlsx", "_expanded.xlsx")
    wb.save(temp_path)
    return temp_path

def write_data_to_template(template_path, output_path, data_df, value_columns, write_columns):
    wb = load_workbook(template_path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    template_df = pd.DataFrame(rows[1:], columns=rows[0])
    point_col = template_df.columns[0]
    point_ids = template_df[point_col].astype(str)
    point_index = {pid: i + 2 for i, pid in enumerate(point_ids) if pid != 'nan'}

    all_value_fields = value_columns  # 这里应已是 fixed_cells + keyword_fields

    for val_col, write_col in zip(all_value_fields, write_columns):
        col_index = column_index_from_string(write_col)
        for _, row in data_df.iterrows():
            full_id = str(row["编号"]).strip()
            if full_id not in point_index:
                print(f"[警告] 模板中找不到编号：{full_id}")
                continue
            value = row.get(val_col, None)
            if pd.notna(value):
                r = point_index[full_id]
                ws.cell(row=r, column=col_index).value = value

    wb.save(output_path)
    print(f"✅ 写入完成，包含固定单元格与偏移字段数据：{output_path}")

# ====== 示例调用入口 ======

if __name__ == "__main__":
    folder = "F:\\样方统计"
    output_path = "F:\\result.xlsx"
    template_path = "F:\\新(模板）2025年6月巴里坤调查点统计.xlsx"
    fixed_cells = ["D2", "F2"]
    keyword_config = {
        "株数": {
            "aliases": ["株数", "株树", "丛数"],
            "offset": (11, 0)
        },
        "面积占比": {
            "aliases": ["面积"],
            "offset": (10, 0)
        },
        "最大株高": {
            "aliases": ["株高"],
            "offset": (10, 0)
        },
        "最大直径": {
            "aliases": ["直径"],
            "offset": (10, 0)
        },
        "土壤含水量": {
            "aliases": ["含水率"],
            "offset": (10, 0)
        },
        "土壤电导率": {
            "aliases": ["电导率"],
            "offset": (10, 0)
        },
        "周长": {
            "aliases": ["周长"],
            "offset": (10, 0)
        }
    }
    keyword_search_range = ("C4", "C100")  # 查找区域
    write_columns = ["AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN", "AP"]
    value_columns = fixed_cells + list(keyword_config.keys())
    df = process_excel_folder(folder, fixed_cells, keyword_config, keyword_search_range)
    expanded_template = preprocess_template_expand_sub_ids(template_path, df)
    write_data_to_template(expanded_template, output_path, df, value_columns, write_columns)